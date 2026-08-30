"""Parses FO_Master_Consolidado.xlsx into clean pandas structures.

This module is the single place that knows about the workbook's layout (which row a
table starts on, which columns it uses, which rows are subtotals). Every other module
— KPI engine, simulator, pending-data scanner, and every Streamlit page — consumes the
`FOData` object this module returns and never touches openpyxl directly. That keeps a
future migration to a different UI (e.g. Next.js/FastAPI) limited to swapping the
presentation layer.
"""

from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import openpyxl
import pandas as pd

from src.excel_recalc import ensure_recalculated

DATA_DIR = Path(__file__).resolve().parent.parent / "data"


# --------------------------------------------------------------------------- #
# Small helpers
# --------------------------------------------------------------------------- #


def _num(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def _rows(ws, first: int, last: int, cols: str = "A:Z"):
    """Yield (row_idx, [values]) for row indices first..last inclusive."""
    col_start, col_end = cols.split(":")
    for row in ws.iter_rows(min_row=first, max_row=last, min_col=openpyxl.utils.column_index_from_string(col_start),
                             max_col=openpyxl.utils.column_index_from_string(col_end)):
        yield row[0].row, [c.value for c in row]


# --------------------------------------------------------------------------- #
# Data containers
# --------------------------------------------------------------------------- #


@dataclass
class LiquidezData:
    detalle: pd.DataFrame
    subtotal_perimetro_completo: float
    subtotal_nucleo: float
    referencia_fo: pd.DataFrame
    referencia_fo_total: float


@dataclass
class InversionesData:
    detalle: pd.DataFrame
    subtotal_perimetro_completo: float
    subtotal_nucleo: float
    referencia_fo: pd.DataFrame
    referencia_fo_total: float


@dataclass
class OtrasPartidasData:
    provisiones: pd.DataFrame
    provisiones_subtotal_memo: float
    cuentas_por_cobrar: pd.DataFrame
    cuentas_por_cobrar_subtotal: float
    otros_menores: pd.DataFrame
    otros_menores_subtotal: float


@dataclass
class FlujoCajaData:
    ingresos: pd.DataFrame  # index=concepto, columns=60 month periods
    egresos: pd.DataFrame
    total_ingresos: pd.Series
    total_egresos: pd.Series
    saldo_mensual: pd.Series
    saldo_acumulado: pd.Series  # reconstructed, see note below
    saldo_acumulado_is_reconstructed: bool
    resumen_anual: pd.DataFrame
    meses: list  # datetime.date, one per column, chronological


@dataclass
class BalanceData:
    activos: pd.DataFrame
    total_activos: float
    pasivos: pd.DataFrame
    total_pasivos: float
    patrimonio_neto: float
    patrimonio_neto_ajustado: float
    activos_contingentes_memo: float
    distribucion_activos: pd.DataFrame
    rendimiento_anual: pd.DataFrame


@dataclass
class FOData:
    fecha_carga: str
    supuestos: dict[str, Any]
    supuestos_raw: pd.DataFrame
    perimetro_nucleo: pd.DataFrame
    perimetro_fuera: pd.DataFrame
    liquidez: LiquidezData
    inversiones: InversionesData
    otras_partidas: OtrasPartidasData
    bienes_raices: pd.DataFrame
    empresas: pd.DataFrame
    pasivos: pd.DataFrame
    total_pasivos: float
    activos_contingentes: pd.DataFrame
    total_activos_contingentes: float
    flujo_caja: FlujoCajaData
    balance: BalanceData
    kpis_raw: dict[str, pd.DataFrame]


# --------------------------------------------------------------------------- #
# Per-sheet parsers
# --------------------------------------------------------------------------- #


def _parse_supuestos(wb) -> tuple[dict, pd.DataFrame]:
    ws = wb["Supuestos"]
    rows = []
    for r, vals in _rows(ws, 4, 17, "A:C"):
        param, valor, nota = vals[0], vals[1], vals[2]
        if param is None:
            continue
        rows.append({"Parámetro": param, "Valor": valor, "Nota / Fuente": nota})
    df = pd.DataFrame(rows)

    def find(substr: str) -> Any:
        match = df[df["Parámetro"].str.contains(substr, case=False, na=False)]
        return match.iloc[0]["Valor"] if not match.empty else None

    supuestos = {
        "valor_uf": _num(find("Valor UF")),
        "tc_usd_clp": _num(find("Tipo de cambio USD")),
        "factor_tasacion": _num(find("Factor Tasación")) or 1.0,
        "inflacion_anual": _num(find("Inflación anual")) or 0.0,
        "retorno_liquidez": _num(find("Retorno esperado anual — Liquidez")) or 0.0,
        "retorno_inversiones": _num(find("Retorno esperado anual — Inversiones")) or 0.0,
        "retorno_bienes_raices": _num(find("Retorno esperado anual — Bienes")) or 0.0,
        "retorno_empresas": _num(find("Retorno esperado anual — Empresas")) or 0.0,
        "tasa_descuento": _num(find("Tasa de descuento")) or 0.0,
    }
    return supuestos, df


def _parse_perimetro(wb) -> tuple[pd.DataFrame, pd.DataFrame]:
    ws = wb["Perimetro_Familiar"]
    nucleo_rows = []
    for r, vals in _rows(ws, 5, 9, "A:D"):
        if vals[0] is None:
            continue
        nucleo_rows.append({"Persona": vals[0], "RUT": vals[1], "Rol": vals[2], "Vinculación societaria": vals[3]})
    fuera_rows = []
    for r, vals in _rows(ws, 13, 15, "A:C"):
        if vals[0] is None:
            continue
        fuera_rows.append({"Persona": vals[0], "RUT": vals[1], "Nota": vals[2]})
    return pd.DataFrame(nucleo_rows), pd.DataFrame(fuera_rows)


def _parse_liquidez_like(ws, detail_first: int, detail_last: int) -> tuple[pd.DataFrame, float, float]:
    rows = []
    subtotal_completo = None
    subtotal_nucleo = None
    for r, vals in _rows(ws, detail_first, detail_last, "A:E"):
        label = vals[0]
        if label is None:
            continue
        if isinstance(label, str) and label.upper().startswith("SUBTOTAL"):
            monto = _num(vals[3])
            if "NÚCLEO" in label.upper() or "NUCLEO" in label.upper():
                subtotal_nucleo = monto
            else:
                subtotal_completo = monto
            continue
        rows.append({
            "Cuenta/Instrumento": label,
            "Titular/Entidad": vals[1],
            "Perímetro": vals[2],
            "Monto (CLP)": _num(vals[3]),
            "Nota": vals[4],
        })
    return pd.DataFrame(rows), subtotal_completo, subtotal_nucleo


def _parse_liquidez(wb) -> LiquidezData:
    ws = wb["Liquidez"]
    detalle, sub_completo, sub_nucleo = _parse_liquidez_like(ws, 5, 33)
    ref_rows = []
    ref_total = 0.0
    for r, vals in _rows(ws, 37, 43, "A:E"):
        if vals[0] is None:
            continue
        if isinstance(vals[0], str) and vals[0].startswith("Total"):
            ref_total = _num(vals[3]) or 0.0
            continue
        ref_rows.append({"Plataforma": vals[0], "Titular": vals[1], "Monto (CLP)": _num(vals[3]), "Nota": vals[4]})
    return LiquidezData(
        detalle=detalle,
        subtotal_perimetro_completo=sub_completo or 0.0,
        subtotal_nucleo=sub_nucleo or 0.0,
        referencia_fo=pd.DataFrame(ref_rows),
        referencia_fo_total=ref_total,
    )


def _parse_inversiones(wb) -> InversionesData:
    ws = wb["Inversiones_Financieras"]
    detalle, sub_completo, sub_nucleo = _parse_liquidez_like(ws, 5, 38)
    ref_rows = []
    ref_total = 0.0
    for r, vals in _rows(ws, 42, 59, "A:D"):
        if vals[0] is None:
            continue
        if isinstance(vals[0], str) and vals[0].startswith("Total"):
            ref_total = _num(vals[2])
            continue
        ref_rows.append({"Institución/Instrumento": vals[0], "Titular": vals[1], "Monto (CLP)": _num(vals[2]), "Nota": vals[3]})
    return InversionesData(
        detalle=detalle,
        subtotal_perimetro_completo=sub_completo or 0.0,
        subtotal_nucleo=sub_nucleo or 0.0,
        referencia_fo=pd.DataFrame(ref_rows),
        referencia_fo_total=ref_total,
    )


def _parse_otras_partidas(wb) -> OtrasPartidasData:
    ws = wb["Otras_Partidas"]

    def block(first: int, last: int) -> tuple[pd.DataFrame, float]:
        rows = []
        subtotal = 0.0
        for r, vals in _rows(ws, first, last, "A:B"):
            if vals[0] is None:
                continue
            if isinstance(vals[0], str) and vals[0].lower().startswith("subtotal"):
                subtotal = _num(vals[1]) or 0.0
                continue
            rows.append({"Concepto": vals[0], "Monto (CLP)": _num(vals[1])})
        return pd.DataFrame(rows), subtotal

    provisiones, provisiones_subtotal = block(6, 10)
    cxc, cxc_subtotal = block(14, 21)
    otros, otros_subtotal = block(25, 31)
    return OtrasPartidasData(
        provisiones=provisiones,
        provisiones_subtotal_memo=provisiones_subtotal,
        cuentas_por_cobrar=cxc,
        cuentas_por_cobrar_subtotal=cxc_subtotal,
        otros_menores=otros,
        otros_menores_subtotal=otros_subtotal,
    )


def _parse_bienes_raices(wb, factor_tasacion: float) -> pd.DataFrame:
    ws = wb["Bienes_Raices"]
    rows = []
    for r, vals in _rows(ws, 5, 38, "A:N"):
        tipo, direccion = vals[1], vals[2]
        if tipo is None and direccion is None:
            continue
        if isinstance(vals[6], str) and vals[6].strip().upper() == "TOTAL":
            continue
        avaluo = _num(vals[7])
        tasacion = _num(vals[8])
        pct = _num(vals[9])
        pct_confirmado = pct is not None
        if pct is None:
            pct = 1.0

        if tasacion is not None:
            valor_balance = tasacion
            fuente = "Tasación directa/comercial"
        elif avaluo is not None:
            valor_balance = avaluo * factor_tasacion
            fuente = "Avalúo Fiscal x factor (proxy)"
        else:
            valor_balance = None
            fuente = "Sin dato (pendiente)"

        valor_atribuible = valor_balance * pct if valor_balance is not None else None

        rows.append({
            "Tipo": tipo,
            "Dirección": direccion,
            "Comuna": vals[3],
            "Rol SII": vals[4],
            "Sup. (m²)": _num(vals[5]),
            "Titular": vals[6],
            "Avalúo Fiscal (CLP)": avaluo,
            "Tasación Comercial (CLP)": tasacion,
            "% Participación": pct,
            "% Participación confirmado": pct_confirmado,
            "Valor Balance (CLP)": valor_balance,
            "Valor Atribuible (CLP)": valor_atribuible,
            "Fuente valor usado": fuente,
        })
    df = pd.DataFrame(rows).reset_index(drop=True)
    df.insert(0, "ID", range(1, len(df) + 1))
    return df


def _parse_empresas(wb) -> pd.DataFrame:
    ws = wb["Empresas"]
    rows = []
    for r, vals in _rows(ws, 5, 38, "A:F"):
        empresa = vals[0]
        if empresa is None:
            continue
        if isinstance(empresa, str) and empresa.upper().startswith("TOTAL"):
            continue
        rut = vals[1]
        pct = _num(vals[2])
        patrimonio = _num(vals[3])
        equity = pct * patrimonio if (pct is not None and patrimonio is not None) else 0.0
        rut_confirmado = isinstance(rut, str) and "pendiente" not in rut.lower() and "verificar" not in rut.lower()
        rows.append({
            "Empresa": empresa,
            "RUT": rut,
            "RUT confirmado": rut_confirmado,
            "% Participación (provisorio)": pct,
            "Patrimonio Contable (CLP)": patrimonio,
            "Equity Value Estimado (CLP)": equity,
            "Nota": vals[5],
        })
    return pd.DataFrame(rows)


def _parse_pasivos(wb) -> tuple[pd.DataFrame, float]:
    ws = wb["Pasivos"]
    rows = []
    total = 0.0
    for r, vals in _rows(ws, 5, 10, "A:E"):
        if vals[0] is None and (isinstance(vals[1], str) and vals[1].upper().startswith("TOTAL")):
            total = _num(vals[3]) or 0.0
            continue
        if vals[1] is None:
            continue
        rows.append({"RUT": vals[0], "Empresa": vals[1], "Comuna": vals[2], "Deuda (CLP)": _num(vals[3]), "Nota": vals[4]})
    return pd.DataFrame(rows), total


def _parse_activos_contingentes(wb) -> tuple[pd.DataFrame, float]:
    ws = wb["Activos_Contingentes"]
    rows = []
    total = 0.0
    for r, vals in _rows(ws, 5, 8, "A:C"):
        if vals[0] is None:
            continue
        if isinstance(vals[0], str) and vals[0].upper().startswith("TOTAL"):
            total = _num(vals[1]) or 0.0
            continue
        rows.append({"Concepto": vals[0], "Monto (CLP)": _num(vals[1]), "Estado / Nota": vals[2]})
    return pd.DataFrame(rows), total


_MONTH_ES = ["ene", "feb", "mar", "abr", "may", "jun", "jul", "ago", "sep", "oct", "nov", "dic"]


def _flujo_caja_month_index() -> list:
    return list(pd.date_range("2026-08-01", periods=60, freq="MS").date)


def _parse_flujo_caja(wb) -> FlujoCajaData:
    ws = wb["Flujo_Caja"]
    meses = _flujo_caja_month_index()
    n = len(meses)
    first_col, last_col = 2, 2 + n - 1  # B..BI

    def month_row(row_idx: int) -> pd.Series:
        vals = [ws.cell(row=row_idx, column=c).value for c in range(first_col, last_col + 1)]
        return pd.Series([_num(v) for v in vals], index=meses)

    ingresos_rows = {}
    for r in range(6, 26):  # includes "Efecto Inflacion" at 25
        label = ws.cell(row=r, column=1).value
        if not label:
            continue
        ingresos_rows[label] = month_row(r)
    total_ingresos = month_row(26)

    egresos_rows = {}
    for r in range(29, 80):  # includes "Efecto Inflacion" at 79
        label = ws.cell(row=r, column=1).value
        if not label:
            continue
        egresos_rows[label] = month_row(r)
    total_egresos = month_row(80)

    saldo_mensual = month_row(82)

    resumen_rows = []
    for r, vals in _rows(ws, 89, 92, "A:F"):
        resumen_rows.append({
            "Horizonte": vals[0],
            "Total Ingresos": _num(vals[1]),
            "Total Egresos": _num(vals[2]),
            "Saldo Mensual": _num(vals[3]),
            "Saldo Acumulado": _num(vals[4]),
            "Activos Líquidos Totales (modelo)": _num(vals[5]),
        })
    resumen_anual = pd.DataFrame(resumen_rows)

    # The monthly "Saldo Acumulado" row in the source workbook is blank (the original
    # model's cumulative series wasn't carried into this master file). We reconstruct
    # it as a running sum of Saldo Mensual, anchored so it agrees exactly with the
    # known truth point at oct-2026 from the "Resumen Anual" table (which also backs
    # the KPI "Peor Saldo Acumulado proyectado"). This is flagged as reconstructed
    # data throughout the UI and in the Datos Pendientes panel.
    running = saldo_mensual.fillna(0).cumsum()
    anchor_label = "Año 1 (oct-2026)"
    anchor_row = resumen_anual[resumen_anual["Horizonte"] == anchor_label]
    if not anchor_row.empty and pd.Timestamp("2026-10-01").date() in running.index:
        anchor_value = anchor_row.iloc[0]["Saldo Acumulado"]
        anchor_running = running.loc[pd.Timestamp("2026-10-01").date()]
        offset = anchor_value - anchor_running
    else:
        offset = 0.0
    saldo_acumulado = running + offset

    ingresos_df = pd.DataFrame(ingresos_rows).T
    egresos_df = pd.DataFrame(egresos_rows).T

    return FlujoCajaData(
        ingresos=ingresos_df,
        egresos=egresos_df,
        total_ingresos=total_ingresos,
        total_egresos=total_egresos,
        saldo_mensual=saldo_mensual,
        saldo_acumulado=saldo_acumulado,
        saldo_acumulado_is_reconstructed=True,
        resumen_anual=resumen_anual,
        meses=meses,
    )


def _parse_balance(wb) -> BalanceData:
    ws = wb["Balance_Consolidado"]
    activos_rows = []
    for r, vals in _rows(ws, 6, 11, "A:D"):
        if vals[0] is None:
            continue
        activos_rows.append({"Categoría": vals[0], "Monto (CLP)": _num(vals[1]), "% del Total": _num(vals[2]), "Fuente": vals[3]})
    total_activos = _num(ws.cell(row=12, column=2).value)

    pasivos_rows = []
    for r, vals in _rows(ws, 16, 16, "A:D"):
        pasivos_rows.append({"Categoría": vals[0], "Monto (CLP)": _num(vals[1])})
    total_pasivos = _num(ws.cell(row=17, column=2).value)

    patrimonio_neto = _num(ws.cell(row=20, column=2).value)
    contingentes_memo = _num(ws.cell(row=21, column=2).value)
    patrimonio_ajustado = _num(ws.cell(row=22, column=2).value)

    distribucion_rows = []
    for r, vals in _rows(ws, 27, 31, "A:C"):
        if vals[0] is None:
            continue
        distribucion_rows.append({"Clase de Activo": vals[0], "Monto (CLP)": _num(vals[1]), "% del Total": _num(vals[2])})

    rendimiento_rows = []
    for r, vals in _rows(ws, 36, 39, "A:D"):
        if vals[0] is None:
            continue
        rendimiento_rows.append({"Clase de Activo": vals[0], "Retorno Anual (supuesto)": _num(vals[1]), "Tipo": vals[2], "Fuente": vals[3]})

    return BalanceData(
        activos=pd.DataFrame(activos_rows),
        total_activos=total_activos,
        pasivos=pd.DataFrame(pasivos_rows),
        total_pasivos=total_pasivos,
        patrimonio_neto=patrimonio_neto,
        patrimonio_neto_ajustado=patrimonio_ajustado,
        activos_contingentes_memo=contingentes_memo,
        distribucion_activos=pd.DataFrame(distribucion_rows),
        rendimiento_anual=pd.DataFrame(rendimiento_rows),
    )


def _parse_kpis(wb) -> dict[str, pd.DataFrame]:
    ws = wb["KPIs"]

    def section(first: int, last: int, cols: str = "A:E") -> pd.DataFrame:
        rows = []
        for r, vals in _rows(ws, first, last, cols):
            if vals[0] is None:
                continue
            rows.append(vals)
        return pd.DataFrame(rows)

    a = section(6, 15)
    a.columns = ["KPI", "Valor", "_c", "_d", "Fórmula / Fuente"]
    a = a[["KPI", "Valor", "Fórmula / Fuente"]]

    b = section(19, 21)
    b.columns = ["KPI", "Valor", "_c", "_d", "Fórmula / Fuente"]
    b = b[["KPI", "Valor", "Fórmula / Fuente"]]

    c = section(25, 28)
    c.columns = ["KPI", "1 año", "5 años", "10 años", "Fórmula / Fuente"]

    d = section(32, 34)
    d.columns = ["KPI", "Valor", "_c", "_d", "Fórmula / Fuente"]
    d = d[["KPI", "Valor", "Fórmula / Fuente"]]

    return {"balance_estructura": a, "liquidez_cobertura": b, "rentabilidad_crecimiento": c, "riesgo_concentracion": d}


# --------------------------------------------------------------------------- #
# Public entry point
# --------------------------------------------------------------------------- #


def load_fo_data(data_dir: Path | None = None, force_recalc: bool = False) -> FOData:
    data_dir = data_dir or DATA_DIR
    recalc_path = ensure_recalculated(data_dir, force=force_recalc)
    wb = openpyxl.load_workbook(recalc_path, data_only=True)

    supuestos, supuestos_raw = _parse_supuestos(wb)
    perimetro_nucleo, perimetro_fuera = _parse_perimetro(wb)
    liquidez = _parse_liquidez(wb)
    inversiones = _parse_inversiones(wb)
    otras_partidas = _parse_otras_partidas(wb)
    bienes_raices = _parse_bienes_raices(wb, supuestos["factor_tasacion"])
    empresas = _parse_empresas(wb)
    pasivos, total_pasivos = _parse_pasivos(wb)
    activos_contingentes, total_activos_contingentes = _parse_activos_contingentes(wb)
    flujo_caja = _parse_flujo_caja(wb)
    balance = _parse_balance(wb)
    kpis_raw = _parse_kpis(wb)

    import datetime

    fecha_carga = datetime.datetime.fromtimestamp(recalc_path.stat().st_mtime).strftime("%Y-%m-%d %H:%M")

    return FOData(
        fecha_carga=fecha_carga,
        supuestos=supuestos,
        supuestos_raw=supuestos_raw,
        perimetro_nucleo=perimetro_nucleo,
        perimetro_fuera=perimetro_fuera,
        liquidez=liquidez,
        inversiones=inversiones,
        otras_partidas=otras_partidas,
        bienes_raices=bienes_raices,
        empresas=empresas,
        pasivos=pasivos,
        total_pasivos=total_pasivos,
        activos_contingentes=activos_contingentes,
        total_activos_contingentes=total_activos_contingentes,
        flujo_caja=flujo_caja,
        balance=balance,
        kpis_raw=kpis_raw,
    )
